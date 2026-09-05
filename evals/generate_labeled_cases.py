"""Generates evals/labeled_cases.xlsx: the eval harness's hand-labeled input sheet (build
order item 9). One row per case, headers plus dropdown data validation on expected_track and
expected_level so a typo can't silently produce an uncomputable case (evals/run_evals.py
would otherwise have to guess what "L4 " or "l4" or "Fellow" means).

This script is a one-time (or deliberate re-run with --force) bootstrap, not something
evals/run_evals.py ever calls -- once a comp professional starts hand-labeling the sheet,
regenerating it would destroy that work. Run again over an existing file and it refuses
unless told --force, the same "don't overwrite what a human is editing by hand" discipline
evals/run_evals.py applies to this same file from the other direction (it only ever reads
labeled_cases.xlsx, never writes it).

Pre-filled rows: the 5 cases that already have a reviewed baseline (scripts/
level_five_jobs_via_graph.py's original CASES, carried over from the retired
evals/labeled_cases.jsonl) plus the 15 real Nyx census rows (NYX-001..NYX-015, deterministic,
first-15-in-order -- data/generate.py's build_nyx_census) with every *expected_* column left
blank, awaiting a comp professional's actual label (CLAUDE.md: leveling-decision domain calls
are the user's, never invented here). All 15 share one source_headcount/source_stage/
source_type -- data/parquet/acquisition_context.parquet is one deal-level record for the
whole census, not per employee.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import get_args

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from agents.schemas import LevelCode

OUTPUT_PATH = Path(__file__).resolve().parent / "labeled_cases.xlsx"
CENSUS_PATH = Path(__file__).resolve().parent.parent / "data" / "parquet" / "nyx_census.xlsx"
ACQUISITION_CONTEXT_PATH = Path(__file__).resolve().parent.parent / "data" / "parquet" / "acquisition_context.parquet"

COLUMNS = [
    "case_id", "source", "role_summary", "source_headcount", "source_stage", "source_type",
    "expected_track", "expected_level", "expected_escalate", "rule_under_test", "label_notes",
]
TRACK_OPTIONS = ["IC", "MGR"]
LEVEL_OPTIONS = list(get_args(LevelCode))

# The 5 already-reviewed baselines, carried over verbatim from the retired
# evals/labeled_cases.jsonl (scripts/level_five_jobs_via_graph.py's original CASES) --
# structural results already treated as ground truth, not re-derived here.
_BASELINE_ROWS = [
    {
        "case_id": "case-01",
        "source": "synthetic",
        "role_summary": (
            "Physical Design Engineer. Owns place-and-route and timing closure for a "
            "subsystem within our next-generation SoC, working independently across the "
            "full development cycle from RTL handoff through tapeout. Sets their own "
            "methodology for the hardest blocks and is consulted by the architecture team "
            "on physical-implementability tradeoffs rather than being directed. Informally "
            "mentors two junior engineers. Influences physical design and timing decisions "
            "within the Physical Design function; not yet driving strategy beyond it. Six "
            "years of related experience."
        ),
        "source_headcount": None, "source_stage": None, "source_type": None,
        "expected_track": "IC", "expected_level": "L4", "expected_escalate": "N",
        "rule_under_test": None,
        "label_notes": "1. Internal IC -- Physical Design",
    },
    {
        "case_id": "case-02",
        "source": "synthetic",
        "role_summary": (
            "Engineering Manager, Firmware. Leads a team of six embedded software "
            "engineers building device driver and RTOS integration work for our sensor "
            "platform. Reviews team output, sets sprint priorities, and represents the "
            "team in cross-functional program reviews. Owns the team's near-term roadmap "
            "and works with product management on scope tradeoffs. No budget authority "
            "beyond headcount requisitions. All six direct reports are individual "
            "contributors."
        ),
        "source_headcount": None, "source_stage": None, "source_type": None,
        "expected_track": "MGR", "expected_level": "M3", "expected_escalate": "N",
        "rule_under_test": None,
        "label_notes": "2. Internal manager -- Embedded Firmware",
    },
    {
        "case_id": "case-03",
        "source": "synthetic",
        "role_summary": (
            "Director of Analog Design. Owned the RF transceiver block design end-to-end "
            "across two tapeouts for our flagship product, from architecture through "
            "characterization. Worked independently with minimal oversight from the VP of "
            "Engineering, who reviewed outcomes rather than approach. Consulted by the "
            "layout and test teams on design-for-test tradeoffs. Represents analog design "
            "in customer technical reviews. Relies heavily on a shared central CAD and "
            "methodology team for tooling and flow support. No direct reports."
        ),
        "source_headcount": 45, "source_stage": "growth", "source_type": "whole company",
        "expected_track": "IC", "expected_level": "L4", "expected_escalate": "Y",
        "rule_under_test": "section 6 rule 3: platform dependency must be assessed for carve-outs",
        "label_notes": '3. Acquired -- "Director of Analog Design"',
    },
    {
        "case_id": "case-04",
        "source": "adversarial",
        "role_summary": (
            "VP of Engineering. Leads all engineering at a 40-person company, with 8 "
            "direct reports, all individual contributors across firmware and hardware "
            "bring-up. Sets sprint priorities and reviews team output on the company's "
            "single embedded product line. Represents engineering in weekly leadership "
            "standups but does not set overall company technical strategy -- that is set "
            "jointly by the two co-founders. No budget authority beyond headcount "
            "requisitions; equipment and vendor spend is approved by the CFO. Problems are "
            "scoped within the current product's firmware and integration issues; the "
            "broader roadmap is set elsewhere."
        ),
        "source_headcount": 40, "source_stage": "growth", "source_type": "whole company",
        "expected_track": "MGR", "expected_level": "M3", "expected_escalate": "N",
        "rule_under_test": "rule 6: title in the source document is evidence, not input",
        "label_notes": '4. Adversarial -- inflated title ("VP of Engineering")',
    },
    {
        "case_id": "case-05",
        "source": "adversarial",
        "role_summary": (
            "Individual contributor with fifteen years focused exclusively on static "
            "timing analysis and timing closure methodology for signoff. Regarded "
            "internally as the final authority on timing closure across the company -- "
            "every product team escalates unresolved timing issues here, and this person "
            "sets the internal timing closure methodology and sign-off criteria for all "
            "tapeouts company-wide, influencing the timing budget across every business "
            "unit. Works with full autonomy, setting direction for the timing closure "
            "domain without oversight. Has never worked outside static timing analysis -- "
            "no experience in place & route, verification, or any adjacent discipline. No "
            "patents, publications, standards body participation, or external industry "
            "recognition."
        ),
        "source_headcount": None, "source_stage": None, "source_type": None,
        "expected_track": "IC", "expected_level": "L5", "expected_escalate": "N",
        "rule_under_test": "rule 3: deep-but-narrow does not reach L6",
        "label_notes": "5. Adversarial -- deep-but-narrow senior IC",
    },
]


def _census_rows() -> list[dict]:
    census = pd.read_excel(CENSUS_PATH).head(15)
    ctx = pd.read_parquet(ACQUISITION_CONTEXT_PATH).iloc[0]

    rows = []
    for i, (_, row) in enumerate(census.iterrows(), start=6):
        rows.append(
            {
                "case_id": f"case-{i:02d}",
                "source": "census",
                "role_summary": row["Role Summary"],
                "source_headcount": int(ctx.source_headcount),
                "source_stage": ctx.source_stage,
                "source_type": ctx.source_type,
                "expected_track": None,
                "expected_level": None,
                "expected_escalate": None,
                "rule_under_test": None,
                "label_notes": f"{row['Emp ID']} -- {row['Job Title']}",
            }
        )
    return rows


def build_rows() -> list[dict]:
    return _BASELINE_ROWS + _census_rows()


def write_workbook(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    ws.append(COLUMNS)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(col) for col in COLUMNS])

    widths = {
        "case_id": 10, "source": 12, "role_summary": 70, "source_headcount": 16,
        "source_stage": 16, "source_type": 16, "expected_track": 14, "expected_level": 14,
        "expected_escalate": 16, "rule_under_test": 46, "label_notes": 40,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[col]

    # Data validation on the two columns a typo would otherwise turn into a silently
    # unscoreable row -- evals/run_evals.py compares expected_level/expected_track against
    # the agent's output by exact string match, so "l4" or "L4 " would just never match
    # anything rather than raising anywhere obvious. Applied through row 1000 so a comp
    # professional adding cases beyond the initial batch keeps the same protection.
    last_row = 1000
    track_col = get_column_letter(COLUMNS.index("expected_track") + 1)
    level_col = get_column_letter(COLUMNS.index("expected_level") + 1)

    track_validation = DataValidation(
        type="list", formula1=f'"{",".join(TRACK_OPTIONS)}"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid track", error="Must be IC or MGR.",
    )
    level_validation = DataValidation(
        type="list", formula1=f'"{",".join(LEVEL_OPTIONS)}"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid level",
        error=f"Must be one of: {', '.join(LEVEL_OPTIONS)}.",
    )
    ws.add_data_validation(track_validation)
    ws.add_data_validation(level_validation)
    track_validation.add(f"{track_col}2:{track_col}{last_row}")
    level_validation.add(f"{level_col}2:{level_col}{last_row}")

    wb.save(path)


def main(force: bool) -> None:
    if OUTPUT_PATH.exists() and not force:
        print(
            f"{OUTPUT_PATH} already exists -- refusing to overwrite hand-labeled work. "
            "Pass --force if you really mean to regenerate it from scratch (this discards "
            "any labels or cases you've added)."
        )
        return
    write_workbook(build_rows(), OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({len(build_rows())} rows).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="overwrite an existing labeled_cases.xlsx")
    args = parser.parse_args()
    main(args.force)
